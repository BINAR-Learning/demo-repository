from fastapi import FastAPI, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
import logging
from datetime import datetime
from typing import List
import uvicorn
from dotenv import load_dotenv

from models import SWOTRequest, SWOTResponse, AnalysisHistory
from services.ai_service import SimpleAIService
from services.db_service import SimpleDBService

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize FastAPI app
app = FastAPI(
    title="Simple AI SWOT Analyzer",
    description="AI-powered SWOT analysis using Google Gemini",
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc"
)

# Add CORS middleware
app.add_middleware(
   
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize services
try:
    ai_service = SimpleAIService()
    db_service = SimpleDBService()
    logger.info("Services initialized successfully")
except Exception as e:
    logger.error(f"Failed to initialize services: {e}")
    raise


@app.get("/")
async def root():
    """Root endpoint with API information"""
    return {
        "message": "Simple AI SWOT Analyzer API",
        "version": "1.0.0",
        "status": "active",
        "timestamp": datetime.now().isoformat(),
        "docs": "/docs",
        "endpoints": {
            "analyze": "POST /analyze",
            "history": "GET /history",
            "stats": "GET /stats",
            "health": "GET /health"
        }
    }


@app.get("/health")
async def health_check():
    """Health check endpoint"""
    try:
        # Test AI service
        ai_status = "ok" if ai_service.api_key else "error"

        # Test DB service
        db_stats = db_service.get_statistics()
        db_status = "ok" if "error" not in db_stats else "error"

        return {
            "status": "healthy" if ai_status == "ok" and db_status == "ok" else "unhealthy",
            "timestamp": datetime.now().isoformat(),
            "services": {
                "ai_service": ai_status,
                "db_service": db_status
            },
            "database": db_stats
        }
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        raise HTTPException(status_code=500, detail=f"Health check failed: {str(e)}")


@app.post("/analyze", response_model=SWOTResponse)
async def analyze_business(request: SWOTRequest,
                          background_tasks: BackgroundTasks):
    """
    Analyze business and generate SWOT analysis using AI

    - **business_description**: Description of the business to analyze
    - **company_name**: Optional company name
    """
    try:
        logger.info(f"Starting SWOT analysis for: {request.company_name or 'unnamed business'}")

        # Generate SWOT analysis using AI
        swot_data = ai_service.generate_swot(
            business_description=request.business_description,
            company_name=request.company_name
        )

        # Create response object
        response = SWOTResponse(
            strengths=swot_data["strengths"],
            weaknesses=swot_data["weaknesses"],
            opportunities=swot_data["opportunities"],
            threats=swot_data["threats"],
            company_name=request.company_name
        )

        # Save to database in background
        background_tasks.add_task(save_analysis_to_db, request, response)

        logger.info("SWOT analysis completed successfully")
        return response

    except Exception as e:
        logger.error(f"Error in analyze_business: {e}")
        raise HTTPException(status_code=500,
                          detail=f"Analysis failed: {str(e)}")


async def save_analysis_to_db(request: SWOTRequest, response: SWOTResponse):
    """Background task to save analysis to database"""
    try:
        analysis_id = db_service.save_analysis(request, response)
        logger.info(f"Analysis saved to database with ID: {analysis_id}")
    except Exception as e:
        logger.error(f"Failed to save analysis to database: {e}")


@app.get("/history", response_model=List[AnalysisHistory])
async def get_analysis_history(limit: int = 10):
    """
    Get recent SWOT analysis history

    - **limit**: Maximum number of results to return (default: 10)
    """
    try:
        if limit < 1 or limit > 100:
            raise HTTPException(status_code=400, detail="Limit must be between 1 and 100")

        history = db_service.get_recent_analyses(limit=limit)
        logger.info(f"Retrieved {len(history)} analysis records")
        return history

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error retrieving history: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to retrieve history: {str(e)}")


@app.get("/history/{analysis_id}")
async def get_analysis_by_id(analysis_id: str):
    """
    Get specific analysis by ID

    - **analysis_id**: UUID of the analysis to retrieve
    """
    try:
        analysis = db_service.get_analysis_by_id(analysis_id)
        if not analysis:
            raise HTTPException(status_code=404, detail="Analysis not found")

        return analysis

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error retrieving analysis {analysis_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to retrieve analysis: {str(e)}")


@app.get("/stats")
async def get_statistics():
    """Get database and usage statistics"""
    try:
        stats = db_service.get_statistics()
        return {
            "api_stats": {
                "service_status": "active",
                "uptime": "running",
                "timestamp": datetime.now().isoformat()
            },
            "database_stats": stats
        }
    except Exception as e:
        logger.error(f"Error getting statistics: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get statistics: {str(e)}")


@app.post("/export/{analysis_id}")
async def export_analysis_to_excel(analysis_id: str):
    """
    Export specific analysis to Excel format

    - **analysis_id**: UUID of the analysis to export
    """
    try:
        # Get analysis data
        analysis = db_service.get_analysis_by_id(analysis_id)
        if not analysis:
            raise HTTPException(status_code=404, detail="Analysis not found")

        # Create Excel file
        excel_file = create_excel_export(analysis)

        return FileResponse(
            path=excel_file,
            filename=f"swot_analysis_{analysis_id[:8]}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting analysis {analysis_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to export analysis: {str(e)}")


def create_excel_export(analysis: AnalysisHistory) -> str:
    """Create Excel file from analysis data"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "SWOT Analysis"

        # Header styling
        header_font = Font(bold=True, size=14)
        category_font = Font(bold=True, size=12)
        category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = f"SWOT Analysis - {analysis.request.company_name or 'Business Analysis'}"
        ws['A1'].font = header_font
        ws['A1'].alignment = Alignment(horizontal='center')

        # Business description
        ws.merge_cells('A2:D2')
        ws['A2'] = f"Business: {analysis.request.business_description}"
        ws['A2'].alignment = Alignment(wrap_text=True)

        # Date
        ws.merge_cells('A3:D3')
        ws['A3'] = f"Analysis Date: {analysis.timestamp}"

        # SWOT sections
        row = 5
        categories = [
            ("STRENGTHS", analysis.response.strengths),
            ("WEAKNESSES", analysis.response.weaknesses),
            ("OPPORTUNITIES", analysis.response.opportunities),
            ("THREATS", analysis.response.threats)
        ]

        for category, items in categories:
            # Category header
            ws[f'A{row}'] = category
            ws[f'A{row}'].font = category_font
            ws[f'A{row}'].fill = category_fill
            row += 1

            # Items
            for i, item in enumerate(items, 1):
                ws[f'A{row}'] = f"{i}."
                ws[f'B{row}'] = item
                ws[f'B{row}'].alignment = Alignment(wrap_text=True)
                row += 1

            row += 1  # Empty row between categories

        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        # Save file
        filename = f"swot_analysis_{analysis.id[:8]}.xlsx"
        wb.save(filename)
        return filename

    except Exception as e:
        logger.error(f"Error creating Excel export: {e}")
        raise


if __name__ == "__main__":
    # Run the application
    uvicorn.run(
        app,
        host="0.0.0.0",
        port=8000,
        log_level="info"
    )
